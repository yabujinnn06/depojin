from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage, ImageDraw, ImageFont
from sqlalchemy import select, func, and_
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import SayimOturumu, Stok, Seri, User
from ..auth import current_user

router = APIRouter(prefix="/api/export", tags=["export"])


DEEP = "FF0F2A44"
ACCENT = "FFBF6F34"
ACCENT_BG = "FFFFEAD6"
GOOD = "FF5FBE7A"
GOOD_BG = "FFE6F4DC"
WARN_BG = "FFFCE5CB"
BAD = "FFDC5A5A"
BAD_BG = "FFFCE0E0"
CREAM = "FFF4F0E8"
EDGE = "FFC6BDAC"


def _border(thin: bool = True) -> Border:
    s = Side(style="thin", color=EDGE) if thin else Side(style="medium", color=EDGE)
    return Border(left=s, right=s, top=s, bottom=s)


_FONT_YOLLARI = [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
    "/Library/Fonts/Arial Bold.ttf",
    "C:/Windows/Fonts/arialbd.ttf",
]


def _font(boyut: int):
    for p in _FONT_YOLLARI:
        try:
            return ImageFont.truetype(p, boyut)
        except Exception:
            continue
    return ImageFont.load_default()


_FILIGRAN_CACHE: bytes | None = None


def _filigran_png() -> bytes:
    global _FILIGRAN_CACHE
    if _FILIGRAN_CACHE is not None:
        return _FILIGRAN_CACHE
    W, H = 900, 280
    img = PILImage.new("RGBA", (W, H), (0, 0, 0, 0))
    d = ImageDraw.Draw(img)
    f1 = _font(140)
    f2 = _font(34)
    txt1 = "YABUJIN"
    txt2 = "depojin sayim sistemi"
    try:
        b1 = d.textbbox((0, 0), txt1, font=f1)
        b2 = d.textbbox((0, 0), txt2, font=f2)
        w1, h1 = b1[2] - b1[0], b1[3] - b1[1]
        w2, h2 = b2[2] - b2[0], b2[3] - b2[1]
    except Exception:
        w1 = h1 = w2 = h2 = 0
    d.text(((W - w1) / 2, (H - h1) / 2 - 18), txt1, fill=(191, 111, 52, 28), font=f1)
    d.text(((W - w2) / 2, (H - h1) / 2 + h1 - 8), txt2, fill=(15, 42, 68, 22), font=f2)
    buf = BytesIO()
    img.save(buf, format="PNG")
    _FILIGRAN_CACHE = buf.getvalue()
    return _FILIGRAN_CACHE


def _filigran_ekle(ws, satir_aralik: int = 22) -> None:
    son_sat = ws.max_row or 1
    son_kol = ws.max_column or 8
    if son_sat < 1:
        return
    konum_satirlari = list(range(8, son_sat + satir_aralik, satir_aralik))
    if not konum_satirlari:
        konum_satirlari = [8]
    for sat in konum_satirlari:
        png = BytesIO(_filigran_png())
        xl = XLImage(png)
        xl.width = 520
        xl.height = 160
        kol_anchor = max(0, (son_kol // 2) - 3)
        xl.anchor = f"{get_column_letter(kol_anchor + 1)}{sat}"
        ws.add_image(xl)


def _baslik_yaz(ws, satir, kolonlar, baslik_renk=DEEP, yazi_renk="FFFFFFFF"):
    for col_idx, (etiket, _w) in enumerate(kolonlar, start=1):
        c = ws.cell(row=satir, column=col_idx, value=etiket)
        c.font = Font(bold=True, color=yazi_renk, size=11)
        c.fill = PatternFill("solid", fgColor=baslik_renk)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()


def _kolon_genislikleri(ws, kolonlar):
    for i, (_e, w) in enumerate(kolonlar, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


@router.get("/{oturum_id}/excel")
def export_excel(oturum_id: int, db: Session = Depends(get_db), _: User = Depends(current_user)):
    oturum = db.get(SayimOturumu, oturum_id)
    if not oturum:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Oturum yok")

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Sayim"
    _ozet_sayfasi(db, wb, oturum)
    _sayim_sayfasi(db, ws1, oturum)
    _log_sayfasi(db, wb, oturum)
    for s in wb.worksheets:
        s.sheet_view.showGridLines = False
        _filigran_ekle(s)
    wb.properties.creator = "Yabujin · Depojin"
    wb.properties.company = "Yabujin"
    wb.properties.title = f"Depojin Sayim — {oturum.ad}"
    wb.properties.subject = "Sayim raporu"

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"sayim_{oturum.ad.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _sayim_sayfasi(db: Session, ws, oturum: SayimOturumu) -> None:
    kolonlar = [
        ("Stok Kodu", 14), ("Urun Adi", 32), ("Seri No", 22),
        ("Portal Sayi", 12), ("Sayilan", 12), ("Fark", 9),
        ("Sayim Tarihi", 19), ("Sayan", 16),
        ("Durum", 12), ("Sonradan Eklendi", 11), ("Cikis Notu", 22),
        ("Zimmet Hedefi", 18), ("Zimmet Notu", 22), ("Notlar", 24),
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(kolonlar))
    c0 = ws.cell(row=1, column=1, value=f"SAYIM RAPORU — {oturum.ad}")
    c0.font = Font(bold=True, color="FFFFFFFF", size=14)
    c0.fill = PatternFill("solid", fgColor=DEEP)
    c0.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    meta = f"Lokasyon: {oturum.lokasyon or '-'}   Durum: {oturum.durum}   Baslangic: {oturum.baslangic.strftime('%d.%m.%Y %H:%M')}"
    if oturum.bitis:
        meta += f"   Bitis: {oturum.bitis.strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(kolonlar))
    c1 = ws.cell(row=2, column=1, value=meta)
    c1.font = Font(italic=True, color="FF555555", size=10)
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    _baslik_yaz(ws, 3, kolonlar)
    _kolon_genislikleri(ws, kolonlar)
    ws.row_dimensions[3].height = 32
    ws.freeze_panes = "D4"

    satir = 4
    stoklar = db.execute(
        select(Stok).where(Stok.oturum_id == oturum.id).order_by(Stok.stok_kodu)
    ).scalars().all()

    for stok in stoklar:
        toplam = len(stok.seriler)
        sayilan_seriler = [s for s in stok.seriler if s.sayildi]
        sayilan = len(sayilan_seriler)
        fark = sayilan - stok.portal_sayim

        baslik_renk = ACCENT_BG if stok.sonradan_eklendi else CREAM
        for col in range(1, len(kolonlar) + 1):
            c = ws.cell(row=satir, column=col)
            c.fill = PatternFill("solid", fgColor=baslik_renk)
            c.border = _border()
            c.font = Font(bold=True, size=11)

        ws.cell(row=satir, column=1, value=stok.stok_kodu)
        ws.cell(row=satir, column=2, value=stok.urun_adi)
        ws.cell(row=satir, column=4, value=stok.portal_sayim)
        ws.cell(row=satir, column=5, value=sayilan)
        farkc = ws.cell(row=satir, column=6, value=fark)
        if fark == 0:
            farkc.fill = PatternFill("solid", fgColor=GOOD_BG); farkc.font = Font(bold=True, color="FF2E7D32")
        elif fark > 0:
            farkc.fill = PatternFill("solid", fgColor=WARN_BG); farkc.font = Font(bold=True, color="FFA9521A")
        else:
            farkc.fill = PatternFill("solid", fgColor=BAD_BG); farkc.font = Font(bold=True, color="FF8A2222")

        if stok.sonradan_eklendi:
            ws.cell(row=satir, column=10, value="EVET")
        satir += 1

        for seri in sorted(stok.seriler, key=lambda x: x.seri_no):
            sayan_ad = None
            if seri.sayan_id:
                u = db.get(User, seri.sayan_id)
                sayan_ad = u.ad if u else None
            zimmet_ad = None
            if seri.zimmet_kullanici_id:
                u = db.get(User, seri.zimmet_kullanici_id)
                zimmet_ad = u.ad if u else None

            durum = "Sayilmadi"
            durum_bg = None
            if seri.cikis_zaman:
                durum = "Cikis"; durum_bg = BAD_BG
            elif seri.zimmet_kullanici_id:
                durum = "Zimmet"; durum_bg = WARN_BG
            elif seri.sayildi:
                durum = "Sayildi"; durum_bg = GOOD_BG

            ws.cell(row=satir, column=3, value=seri.seri_no)
            ws.cell(row=satir, column=7,
                    value=seri.sayim_tarihi.strftime("%d.%m.%Y %H:%M:%S") if seri.sayim_tarihi else "")
            ws.cell(row=satir, column=8, value=sayan_ad or "")
            dc = ws.cell(row=satir, column=9, value=durum)
            if durum_bg:
                dc.fill = PatternFill("solid", fgColor=durum_bg)
                dc.font = Font(bold=True, size=10)
            if seri.sonradan_eklendi:
                ec = ws.cell(row=satir, column=10, value="EVET")
                ec.fill = PatternFill("solid", fgColor=ACCENT_BG)
                ec.font = Font(bold=True, color=ACCENT)
            ws.cell(row=satir, column=11, value=seri.cikis_notu or "")
            ws.cell(row=satir, column=12, value=zimmet_ad or "")
            ws.cell(row=satir, column=13, value=seri.zimmet_notu or "")
            ws.cell(row=satir, column=14, value=seri.notlar or "")

            for col in range(1, len(kolonlar) + 1):
                cc = ws.cell(row=satir, column=col)
                cc.border = _border()
                if col == 3 and seri.sonradan_eklendi:
                    cc.font = Font(bold=True, italic=True, color=ACCENT)
            satir += 1


def _ozet_sayfasi(db: Session, wb: Workbook, oturum: SayimOturumu) -> None:
    ws = wb.create_sheet("Ozet", 0)

    kolonlar = [
        ("Stok Kodu", 14), ("Urun Adi", 36), ("Portal Sayi", 12),
        ("Sayilan", 12), ("Fark", 10), ("Yuzde", 10),
        ("Sonradan Eklendi", 13), ("Cikis", 8), ("Zimmet", 8),
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(kolonlar))
    c0 = ws.cell(row=1, column=1, value=f"OZET — {oturum.ad}")
    c0.font = Font(bold=True, color="FFFFFFFF", size=14)
    c0.fill = PatternFill("solid", fgColor=DEEP)
    c0.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    stoklar = db.execute(
        select(Stok).where(Stok.oturum_id == oturum.id).order_by(Stok.stok_kodu)
    ).scalars().all()
    toplam_portal = sum(s.portal_sayim for s in stoklar)
    toplam_sayilan = sum(1 for s in stoklar for x in s.seriler if x.sayildi)
    toplam_seri = sum(len(s.seriler) for s in stoklar)
    toplam_cikis = sum(1 for s in stoklar for x in s.seriler if x.cikis_zaman)
    toplam_zimmet = sum(1 for s in stoklar for x in s.seriler if x.zimmet_kullanici_id)
    sonradan_seri = sum(1 for s in stoklar for x in s.seriler if x.sonradan_eklendi)
    sonradan_stok = sum(1 for s in stoklar if s.sonradan_eklendi)

    metrikler = [
        ("Stok Kalemi", len(stoklar), DEEP),
        ("Toplam Seri", toplam_seri, DEEP),
        ("Sayilan", toplam_sayilan, GOOD),
        ("Kalan", toplam_seri - toplam_sayilan, ACCENT),
        ("Portalda Toplam", toplam_portal, DEEP),
        ("Portal Fark", toplam_sayilan - toplam_portal, ACCENT),
        ("Sonradan Eklenen Stok", sonradan_stok, ACCENT),
        ("Sonradan Eklenen Seri", sonradan_seri, ACCENT),
        ("Cikis", toplam_cikis, BAD),
        ("Zimmet", toplam_zimmet, ACCENT),
    ]

    for i, (etiket, deger, renk) in enumerate(metrikler):
        col = (i % 5) * 2 + 1
        sat = 3 + (i // 5) * 3
        et = ws.cell(row=sat, column=col, value=etiket)
        et.font = Font(bold=True, size=9, color="FF777777")
        et.alignment = Alignment(horizontal="left")
        dg = ws.cell(row=sat + 1, column=col, value=deger)
        dg.font = Font(bold=True, size=18, color=renk)
        dg.alignment = Alignment(horizontal="left")
        ws.merge_cells(start_row=sat, start_column=col, end_row=sat, end_column=col + 1)
        ws.merge_cells(start_row=sat + 1, start_column=col, end_row=sat + 1, end_column=col + 1)

    bas_satir = 10
    _baslik_yaz(ws, bas_satir, kolonlar)
    _kolon_genislikleri(ws, kolonlar)
    ws.row_dimensions[bas_satir].height = 28
    ws.freeze_panes = f"A{bas_satir + 1}"

    sat = bas_satir + 1
    sirali = sorted(
        stoklar,
        key=lambda s: abs(sum(1 for x in s.seriler if x.sayildi) - s.portal_sayim),
        reverse=True,
    )
    for stok in sirali:
        sayilan = sum(1 for x in stok.seriler if x.sayildi)
        cikis = sum(1 for x in stok.seriler if x.cikis_zaman)
        zimmet = sum(1 for x in stok.seriler if x.zimmet_kullanici_id)
        fark = sayilan - stok.portal_sayim
        yuzde = round(sayilan / stok.portal_sayim * 100, 1) if stok.portal_sayim > 0 else 0

        ws.cell(row=sat, column=1, value=stok.stok_kodu)
        ws.cell(row=sat, column=2, value=stok.urun_adi)
        ws.cell(row=sat, column=3, value=stok.portal_sayim)
        ws.cell(row=sat, column=4, value=sayilan)
        farkc = ws.cell(row=sat, column=5, value=fark)
        ws.cell(row=sat, column=6, value=f"{yuzde}%")
        if stok.sonradan_eklendi:
            ec = ws.cell(row=sat, column=7, value="EVET")
            ec.fill = PatternFill("solid", fgColor=ACCENT_BG)
            ec.font = Font(bold=True, color=ACCENT)
        ws.cell(row=sat, column=8, value=cikis or "")
        ws.cell(row=sat, column=9, value=zimmet or "")

        if fark == 0:
            farkc.fill = PatternFill("solid", fgColor=GOOD_BG); farkc.font = Font(bold=True, color="FF2E7D32")
        elif fark > 0:
            farkc.fill = PatternFill("solid", fgColor=WARN_BG); farkc.font = Font(bold=True, color="FFA9521A")
        else:
            farkc.fill = PatternFill("solid", fgColor=BAD_BG); farkc.font = Font(bold=True, color="FF8A2222")

        for col in range(1, len(kolonlar) + 1):
            ws.cell(row=sat, column=col).border = _border()
        sat += 1


def _log_sayfasi(db: Session, wb: Workbook, oturum: SayimOturumu) -> None:
    from ..models import TaramaLog
    ws = wb.create_sheet("Tarama Log")

    kolonlar = [
        ("Zaman", 19), ("Kullanici", 16), ("Seri Giris", 22),
        ("Durum", 12), ("Stok Kodu", 14), ("Urun Adi", 32), ("Aciklama", 30),
    ]
    _baslik_yaz(ws, 1, kolonlar)
    _kolon_genislikleri(ws, kolonlar)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    rows = db.execute(
        select(TaramaLog, User.ad)
        .outerjoin(User, User.id == TaramaLog.kullanici_id)
        .where(TaramaLog.oturum_id == oturum.id)
        .order_by(TaramaLog.zaman.asc())
    ).all()
    durum_renk = {
        "basarili": GOOD_BG, "mukerrer": WARN_BG, "bulunamadi": BAD_BG, "cakisma": WARN_BG,
    }
    sat = 2
    for log, ad in rows:
        ws.cell(row=sat, column=1, value=log.zaman.strftime("%d.%m.%Y %H:%M:%S"))
        ws.cell(row=sat, column=2, value=ad or "")
        ws.cell(row=sat, column=3, value=log.seri_giris)
        dc = ws.cell(row=sat, column=4, value=log.durum)
        bg = durum_renk.get(log.durum)
        if bg:
            dc.fill = PatternFill("solid", fgColor=bg)
            dc.font = Font(bold=True)
        ws.cell(row=sat, column=5, value=log.stok_kodu or "")
        ws.cell(row=sat, column=6, value=log.urun_adi or "")
        ws.cell(row=sat, column=7, value=log.aciklama or "")
        for col in range(1, len(kolonlar) + 1):
            ws.cell(row=sat, column=col).border = _border()
        sat += 1

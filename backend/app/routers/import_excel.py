from datetime import datetime
from io import BytesIO
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import SayimOturumu, Stok, Seri, User
from ..auth import require_admin
from ..utils import (
    normalize_seri, normalize_stok_kodu, temizle_metin,
    temizle_portal_sayi, is_junk_stok_header, is_system_sheet,
)

router = APIRouter(prefix="/api/import", tags=["import"])


def _veri_sayfasini_sec(wb):
    en_iyi = None
    en_iyi_sayi = 0
    for name in wb.sheetnames:
        if is_system_sheet(name):
            continue
        ws = wb[name]
        if not ws.max_row:
            continue
        sinir = min(ws.max_row, 5000)
        c_sayi = 0
        for row in ws.iter_rows(min_row=2, max_row=sinir, min_col=3, max_col=3, values_only=True):
            v = row[0]
            if v is not None and str(v).strip():
                c_sayi += 1
        if c_sayi > en_iyi_sayi:
            en_iyi = ws
            en_iyi_sayi = c_sayi
    if en_iyi:
        return en_iyi
    for name in wb.sheetnames:
        if not is_system_sheet(name):
            return wb[name]
    return wb[wb.sheetnames[0]]


def _onceki_sayim_durumu(i_col, j_col, k_col) -> tuple[bool, datetime | None, str | None]:
    sayildi = False
    if isinstance(i_col, bool):
        sayildi = i_col
    elif isinstance(i_col, (int, float)):
        sayildi = i_col > 0
    elif i_col is not None and str(i_col).strip():
        sayildi = True
    if not sayildi:
        if j_col is not None and str(j_col).strip():
            sayildi = True
        elif k_col is not None and str(k_col).strip():
            sayildi = True
    if not sayildi:
        return False, None, None
    tarih = j_col if isinstance(j_col, datetime) else None
    kullanici = temizle_metin(k_col) or None
    return True, tarih, kullanici


@router.post("/{oturum_id}/stok-excel")
async def stok_excel(
    oturum_id: int,
    file: UploadFile = File(...),
    devam: bool = Query(False),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    oturum = db.get(SayimOturumu, oturum_id)
    if not oturum:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Oturum yok")

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True, keep_vba=False)
    except Exception as e:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Excel okunamadi: {e}")

    ws = _veri_sayfasini_sec(wb)

    eklenen_stok = 0
    eklenen_seri = 0
    atlanan_junk = 0
    bos_satir = 0
    mukerrer = 0
    onceden_sayilan = 0
    aktif_stok: Stok | None = None

    mevcut_stoklar = {s.stok_kodu: s for s in oturum.stoklar}
    mevcut_seriler: set[tuple[int, str]] = set()
    for s in oturum.stoklar:
        for seri in s.seriler:
            mevcut_seriler.add((s.id, seri.seri_no_norm))

    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        row = list(raw_row)
        while len(row) < 11:
            row.append(None)
        a, b, c, d, e_col, _f, _g, _h, i_col, j_col, k_col = row[:11]

        if all(v is None or (isinstance(v, str) and not v.strip()) for v in (a, b, c, d, e_col)):
            bos_satir += 1
            continue

        if is_junk_stok_header(a):
            atlanan_junk += 1
            continue

        stok_kodu = normalize_stok_kodu(a)
        if stok_kodu:
            urun_adi = temizle_metin(b) or stok_kodu
            portal_sayim = temizle_portal_sayi(d) or temizle_portal_sayi(e_col)

            if stok_kodu in mevcut_stoklar:
                aktif_stok = mevcut_stoklar[stok_kodu]
                if urun_adi:
                    aktif_stok.urun_adi = urun_adi
                if portal_sayim:
                    aktif_stok.portal_sayim = portal_sayim
            else:
                aktif_stok = Stok(
                    oturum_id=oturum_id,
                    stok_kodu=stok_kodu,
                    urun_adi=urun_adi,
                    portal_sayim=portal_sayim,
                )
                db.add(aktif_stok)
                db.flush()
                mevcut_stoklar[stok_kodu] = aktif_stok
                eklenen_stok += 1

        if c is not None and str(c).strip():
            if aktif_stok is None:
                bos_satir += 1
                continue
            seri_raw = temizle_metin(c)
            seri_norm = normalize_seri(seri_raw)
            if not seri_norm:
                bos_satir += 1
                continue
            key = (aktif_stok.id, seri_norm)
            if key in mevcut_seriler:
                mukerrer += 1
                continue
            mevcut_seriler.add(key)

            sayildi = False
            sayim_tarihi = None
            sayan_ad = None
            if devam:
                sayildi, sayim_tarihi, sayan_ad = _onceki_sayim_durumu(i_col, j_col, k_col)
                if sayildi:
                    onceden_sayilan += 1

            db.add(Seri(
                oturum_id=oturum_id,
                stok_id=aktif_stok.id,
                seri_no=seri_raw,
                seri_no_norm=seri_norm,
                sayildi=sayildi,
                sayim_tarihi=sayim_tarihi,
                notlar=(f"Once sayan: {sayan_ad}" if sayan_ad else None),
            ))
            eklenen_seri += 1

    db.commit()
    return {
        "sayfa": ws.title,
        "eklenen_stok": eklenen_stok,
        "eklenen_seri": eklenen_seri,
        "mukerrer_seri": mukerrer,
        "atlanan_junk_header": atlanan_junk,
        "bos_satir": bos_satir,
        "devam_modu": devam,
        "onceden_sayilan_olarak_isaretlenen": onceden_sayilan,
    }
